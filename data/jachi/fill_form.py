# -*- coding: utf-8 -*-
import openpyxl, io, sys, glob
from openpyxl.styles import Alignment, Font, Border, Side
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# 소스: Windows 네이티브 경로(backslash)로 직접 로드 (bash cp/OneDrive 손상 회피)
src = glob.glob(r"D:\Users\gijang\Downloads\자치사무 전수조사 제출 서식.xlsx")
if not src:
    src = glob.glob(r"D:\Users\gijang\Downloads\*전수조사*제출*.xlsx")
print("source:", src)
wb = openpyxl.load_workbook(src[0])
ws = wb["Sheet1"]

#  세무2과 소관(docx '관리책임부서명:세무2과' 확정): 수입증지/제증명 수수료징수/금고지정 규칙/세무조사 운영 규칙
#  → 민원신청서 별지서식은 세무조사 운영 규칙 별지1·2호뿐(납세자 작성), 둘 다 구비서류 없음 → 정비불요
rows = [
 ('규칙','부산광역시 기장군 세무조사 운영 규칙','〔별지 제1호서식〕 지방세 서면조사서','X','해당없음','해당없음','X','서면조사 응답 서식(납세자 작성)·구비서류 없음 → 공동이용 대상 없음, 정비불요'),
 ('규칙','부산광역시 기장군 세무조사 운영 규칙','〔별지 제2호서식〕 공휴일 등 세무조사 실시 요청 및 검토서','X','해당없음','해당없음','X','신청인 인적사항만 기재·구비서류 없음 → 공동이용 대상 없음, 정비불요'),
]

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

for r in range(5, max(ws.max_row, 6) + 1):
    for c in range(1, 11):
        ws.cell(r, c).value = None

for i, (gu, law, form, F, G, H, I, J) in enumerate(rows):
    r = 5 + i
    vals = [i + 1, "세무2과", gu, law, form, F, G, H, I, J]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(r, c, v)
        cell.border = border
        cell.alignment = lft if c in (4, 5, 10) else ctr
        cell.font = Font(name="맑은 고딕", size=9)

widths = {"A": 5, "B": 8, "C": 7, "D": 30, "E": 32, "F": 11, "G": 13, "H": 10, "I": 13, "J": 30}
for col, wd in widths.items():
    ws.column_dimensions[col].width = wd

out = r"D:\Users\gijang\Downloads\자치사무 전수조사_세무2과_초안_v3.xlsx"
wb.save(out)
print("saved:", out, "rows:", len(rows))

# 비동기화 로컬에도 저장 후 읽기검증
scratch = r"C:\Users\gijang\AppData\Local\Temp\claude\C--todo-manual-dashboard-law-dashboard-work\6d4f0b1b-5f80-4282-803e-6f2e7aecea14\scratchpad\se2_draft.xlsx"
wb.save(scratch)
import zipfile
print("scratch zip ok:", zipfile.is_zipfile(scratch))
wb2 = openpyxl.load_workbook(scratch); ws2 = wb2["Sheet1"]
print("verify rows:", ws2.max_row)
for r in range(5, ws2.max_row + 1):
    v = [ws2.cell(r, c).value for c in (1, 4, 5, 9)]
    if v[0] is not None:
        print("  ", v[0], "|", str(v[1])[:22], "|", str(v[2])[:24], "| I=", v[3])
